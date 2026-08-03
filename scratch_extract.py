import json
import openpyxl

# ---- 1. ET 6.1: the 2015 year-end renewable capacity point (year-START for 2016) ----
wb = openpyxl.load_workbook("/tmp/ET_6.1.xlsx", data_only=True)
ws = wb["Annual"]
hdr = {str(ws.cell(7, c).value): c for c in range(2, 40) if ws.cell(7, c).value}
c2015 = hdr["2015"]
onshore15 = float(ws.cell(8, c2015).value)
offshore15 = float(ws.cell(9, c2015).value) + float(ws.cell(10, c2015).value or 0)
solar15 = float(ws.cell(12, c2015).value)
print("ET6.1 2015:", onshore15, offshore15, solar15)
print("  row labels:", ws.cell(8, 1).value, "|", ws.cell(9, 1).value, "|",
      ws.cell(10, 1).value, "|", ws.cell(12, 1).value)

# ---- 2. DUKES 5.7.A: all-generating-companies dispatchable fleet ----
wb2 = openpyxl.load_workbook("/tmp/DUKES_5.7.xlsx", data_only=True)
d = wb2["5.7"]
yrcol = {int(d.cell(7, c).value): c for c in range(3, 40)
         if str(d.cell(7, c).value).isdigit()}
ROWS = {
    "coal_mw": 32, "oil_mw": 33, "gas_mw": 34, "mixed_dual_mw": 35,
    "nuclear_mw": 36, "pumped_hydro_mw": 38, "bioenergy_waste_mw": 43,
    "other_fossil_mw": 44,
}
for k, r in ROWS.items():
    assert d.cell(r, 1).value == "All generating companies", (k, r, d.cell(r, 1).value)
    print(f"  row {r:>3} {k:<20} label={d.cell(r,2).value}")

out = {k: {} for k in ROWS}
tot = {}
for y in range(2015, 2026):
    c = yrcol[y]
    s = 0.0
    for k, r in ROWS.items():
        v = float(d.cell(r, c).value)
        out[k][str(y)] = round(v, 4)
        s += v
    tot[y] = s
print()
mean_1625 = sum(tot[y] for y in range(2016, 2026)) / 10
for y in range(2015, 2026):
    print(f"  {y}  dispatchable={tot[y]:10.1f} MW   coal={out['coal_mw'][str(y)]:9.1f}"
          f"   shape_vs_2016_25_mean={tot[y]/mean_1625:.4f}")
print(f"  window 2016-2025 mean = {mean_1625:.1f} MW")
json.dump({"totals": tot, "series": out}, open("/tmp/disp.json", "w"))
