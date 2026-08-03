"""One-shot ingestion: extend the DUKES capacity series back to 2015 (the year-START
capacity the mean-capacity basis needs for 2016) and write the DUKES 5.7.A dispatchable
fleet series. Scratch script, not committed."""
import json
import openpyxl
from collections import OrderedDict

CAP = "docs/market_research/w1_7_dukes_installed_capacity_annual.json"
DISP = "docs/market_research/w1_7_dukes_dispatchable_capacity_annual.json"

# ---- 1. extend the renewable capacity series back to 2015 (ET 6.1 Annual) ----
wb = openpyxl.load_workbook("/tmp/ET_6.1.xlsx", data_only=True)
ws = wb["Annual"]
hdr = {str(ws.cell(7, c).value): c for c in range(2, 40) if ws.cell(7, c).value}
c = hdr["2015"]
vals15 = {
    "onshore_mw": round(float(ws.cell(8, c).value), 2),
    "offshore_mw": round(float(ws.cell(9, c).value) + float(ws.cell(10, c).value or 0), 2),
    "solar_mw": round(float(ws.cell(12, c).value), 2),
}
cap = json.load(open(CAP), object_pairs_hook=OrderedDict)
for k, v in vals15.items():
    if "2015" in cap[k]:
        raise SystemExit(f"{k} already has 2015 — refusing to overwrite")
    merged = OrderedDict()
    merged["2015"] = v
    merged.update(cap[k])
    cap[k] = merged
cap["_provenance"]["year_2015_addendum"] = (
    "2026-08-03 (W1_7 L2 commissioning-smoothing pass): the 2015 year-end point was added "
    "from the SAME ET 6.1 'Annual' worksheet (column '2015', rows 8/9+10/12). It exists "
    "solely as the year-START capacity for 2016 so the year-MEAN (uniform-commissioning) "
    "basis is computable for every generation year 2016-2025 -- see "
    "docs/market_research/w1_7_commissioning_smoothing_and_restacking.md. The "
    "generation/load-factor table (w1_7_dukes_generation_and_load_factor_annual.json) is "
    "deliberately NOT extended to 2015, so A5/A6 still iterate 2016-2025 exactly as before."
)
json.dump(cap, open(CAP, "w"), indent=2)
print("capacity series extended to 2015:", vals15)

# ---- 2. DUKES 5.7.A dispatchable fleet ----
wb2 = openpyxl.load_workbook("/tmp/DUKES_5.7.xlsx", data_only=True)
d = wb2["5.7"]
yrcol = {int(d.cell(7, x).value): x for x in range(3, 40) if str(d.cell(7, x).value).isdigit()}
ROWS = OrderedDict([
    ("coal_mw", 32), ("oil_mw", 33), ("gas_mw", 34), ("mixed_dual_mw", 35),
    ("nuclear_mw", 36), ("pumped_hydro_mw", 38), ("bioenergy_waste_mw", 43),
    ("other_fossil_mw", 44),
])
LABELS = {}
series = OrderedDict()
for k, r in ROWS.items():
    assert d.cell(r, 1).value == "All generating companies", (k, r)
    LABELS[k] = d.cell(r, 2).value
    series[k] = OrderedDict(
        (str(y), round(float(d.cell(r, yrcol[y]).value), 4)) for y in range(2015, 2026))

out = OrderedDict()
out["_provenance"] = OrderedDict([
    ("source", "DUKES Table 5.7.A 'Capacity by fuel', generator type 'All generating "
               "companies' (rows 32-44 of the '5.7' worksheet), MW"),
    ("url", "https://assets.publishing.service.gov.uk/media/6a6a3622cceb23e8678976bf/DUKES_5.7.xlsx"),
    ("landing_page", "https://www.gov.uk/government/statistics/electricity-chapter-5-"
                     "digest-of-united-kingdom-energy-statistics-dukes"),
    ("fetched", "2026-08-03"),
    ("http_status", 200),
    ("units", "MW, capacity at calendar year end"),
    ("rows_taken", {k: f"row {r} — {LABELS[k]}" for k, r in ROWS.items()}),
    ("basis", "DUKES note 1: grid export capacity where available, else installed "
              "capacity. The de-rating in this table's title applies to WIND, SOLAR and "
              "SMALL-SCALE HYDRO only (note 4) -- none of which are taken here; every "
              "row ingested is a thermal/nuclear/storage row on a grid-export basis."),
    ("rows_deliberately_excluded", [
        "Hydro (natural flow) row 37 — weather-driven renewable, sits on the RENEWABLE "
        "side of the residual-demand identity, not the dispatchable denominator",
        "Onshore wind (39) / Offshore wind (40) / Wave and tidal (41) / Solar (42) — the "
        "renewable side, and de-rated in this table (note 4)",
        "Total rows 31/45/46/47 — aggregates, would double-count",
    ]),
    ("used_by", "sim/renewable_capacity_trend.py — real_dispatchable_capacity_mw(), "
                "dispatchable_shape(), dispatchable_capacity_mw(), real_coal_capacity_by_year(), "
                "A4 (check_no_coal_after_retirement on the REAL series), A10 "
                "(check_dispatchable_fleet_contracts); reaches the merit order via "
                "sim/price_engine.py::system_margin_price(year=...)"),
    ("R10_simplification_geography", "DUKES 5.7 is UNITED KINGDOM; the sim's settlement "
        "data (Elexon) is GB. Northern Ireland (~2 GW, on the SEM not the GB market) is "
        "therefore included in this series. This is a LEVEL error, and it is divided out "
        "by construction: only the SHAPE (each year / the 2016-2025 window mean) is used, "
        "and the calibrated DISPATCHABLE_CAPACITY_MW=35000 level is preserved unchanged."),
    ("R10_simplification_interconnectors", "Interconnector import capacity (~4 GW 2016 -> "
        "~11.7 GW 2025) is part of price_engine's stated definition of the dispatchable "
        "fleet but is NOT in DUKES 5.7 and is NOT ingested here. It grew while the thermal "
        "fleet shrank, so this series OVERSTATES the true contraction of dispatchable "
        "capability. Named, not hidden; closing it needs a NESO interconnector-register pass."),
    ("R13_wall", "BASELINE — real published history, fidelity-to-reality only, never "
                 "tuned for company P&L. Held flat (piecewise-constant) outside 2015-2025; "
                 "the forward window is CURRICULUM and is not authored here."),
    ("year_2015_note", "2015 is carried for the same reason as in the capacity file: it is "
                       "the year-START stock for 2016 under the year-MEAN basis."),
])
out["fuel_mw"] = series
json.dump(out, open(DISP, "w"), indent=2)
print("dispatchable series written:", DISP)
for y in range(2015, 2026):
    print(" ", y, round(sum(series[k][str(y)] for k in series), 1))
